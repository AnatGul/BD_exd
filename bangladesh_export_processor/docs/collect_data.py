# -*- coding: utf-8 -*-
"""
Сбор данных для обучения: JPG (OCR) + Excel (переводы)
Сопоставление: EXD 784-63264471.jpg ↔ EXD 784-63264471 (перевод).xlsx
"""
import os
import re
from typing import Dict, List
from openpyxl import load_workbook

# Сопоставление файлов
FOLDER = r"D:\pythonProject\BD_exd\bangladesh_export_processor\примеры"

# Маппинг номеров полей в названия field_mapper
FIELD_MAP = {
    '1': 'Номер документа',
    '2': 'Наименование экспортера',
    '3': 'Год',
    '5': 'Код офиса экспорта',
    '6': 'Код региона',
    '7': 'Номер коносамента',
    '8': 'Наименование получателя/грузополучателя',
    '14': 'Декларант/Агент',
    '15': 'Адрес декларанта/агента',
    '16': 'Код агента',
    '17': 'Условия поставки',
    '18': 'Код условий поставки',
    '20': 'Код валюты',
    '22': 'Общая стоимость',
    '23': 'Курс валют',
    '24': 'Страна происхождения',
    '25': 'Порт погрузки',
    '27': 'Наименование авиакомпании',
    '28': 'Код авиакомпании',
    '29': 'Наименование банка',
    '30': 'Код банка',
    '31': 'Общая декларируемая стоимость',
    '32': 'Номер места',
    '33': 'Код ТН ВЭД',
    '35': 'Описание товара',
    '37': 'Количество мест (ед)',
    '38': 'Количество мест (ед.изм)',
    '41': 'Номер CRF/EXP',
    '44': 'Сектор и фонд',
}


def get_excel_fields(excel_path: str) -> Dict[str, str]:
    """Читает Excel: A+B+C (C - продолжение B)"""
    fields = {}
    
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(values_only=True):
        field_name = row[0]  # Колонка A - номер и название
        value_b = row[1] if row[1] else ''  # Колонка B
        value_c = row[2] if row[2] else ''  # Колонка C - продолжение
        
        if field_name:
            # Извлечь номер поля (например "2. Грузоотправитель" -> "2")
            match = re.match(r'^(\d+)\.', str(field_name))
            if match:
                field_num = match.group(1)
                # Склеить B + C
                full_value = str(value_b) + str(value_c)
                if full_value.strip() and full_value.strip() != 'None':
                    fields[field_num] = full_value.strip()
    
    return fields


def run_ocr(image_path: str) -> List[str]:
    """Запускает Tesseract OCR на изображении"""
    from src.ocr_reader import OCRReader
    
    reader = OCRReader(use_tesseract=True, use_preprocessing=True)
    results = reader.read_image(image_path)
    return [r['text'] for r in results]


def find_ocr_matches(ocr_text: str, excel_fields: Dict[str, str]) -> Dict[str, List[str]]:
    """Ищет какие значения из Excel присутствуют в OCR тексте"""
    matches = {}
    
    for field_num, value in excel_fields.items():
        # Ищем первые 30 символов значения в OCR
        search_term = value[:30].lower()
        ocr_lower = ocr_text.lower()
        
        if search_term in ocr_lower:
            if field_num not in matches:
                matches[field_num] = []
            matches[field_num].append(value[:50])
    
    return matches


def collect_all_data():
    """Собирает данные со всех файлов"""
    all_data = []
    
    # Найти все JPG файлы
    jpg_files = [f for f in os.listdir(FOLDER) if f.endswith('.jpg')]
    print(f"Найдено {len(jpg_files)} JPG файлов")
    
    for jpg_file in jpg_files:
        base_name = jpg_file.replace('.jpg', '')
        excel_file = f"{base_name} (перевод).xlsx"
        excel_path = os.path.join(FOLDER, excel_file)
        
        if not os.path.exists(excel_path):
            print(f"  Excel не найден: {excel_file}")
            continue
        
        print(f"\n=== {base_name} ===")
        
        # Читать Excel
        excel_fields = get_excel_fields(excel_path)
        print(f"  Excel: {len(excel_fields)} полей")
        
        # Запустить OCR
        jpg_path = os.path.join(FOLDER, jpg_file)
        print(f"  Running OCR...")
        ocr_texts = run_ocr(jpg_path)
        ocr_text = ' '.join(ocr_texts)
        print(f"  OCR: {len(ocr_texts)} элементов, {len(ocr_text)} символов")
        
        # Найти совпадения
        matches = find_ocr_matches(ocr_text, excel_fields)
        print(f"  Совпадений: {len(matches)}")
        
        all_data.append({
            'base_name': base_name,
            'excel_fields': excel_fields,
            'ocr_text': ocr_text,
            'matches': matches,
        })
    
    return all_data


def analyze_patterns(all_data: List[Dict]):
    """Анализирует паттерны из обучающей выборки"""
    print("\n" + "="*60)
    print("АНАЛИЗ ПАТТЕРНОВ")
    print("="*60)
    
    # Собрать все значения для каждого поля
    field_values = {}
    
    for data in all_data:
        for field_num, value in data['excel_fields'].items():
            if field_num not in field_values:
                field_values[field_num] = []
            field_values[field_num].append(value)
    
    # Вывести статистику по полям
    print("\n=== Поля по количеству значений ===")
    for field_num in sorted(field_values.keys(), key=lambda x: int(x)):
        values = field_values[field_num]
        field_name = FIELD_MAP.get(field_num, f'Поле {field_num}')
        
        # Уникальные значения
        unique_vals = list(set(values))[:3]
        
        print(f"\n{field_num}. {field_name} ({len(values)} файлов):")
        for v in unique_vals:
            print(f"   - {v[:60]}")
    
    return field_values


def generate_pattern_suggestions(field_values: Dict[str, List[str]]):
    """Генерирует предложения по паттернам"""
    print("\n" + "="*60)
    print("ПРЕДЛОЖЕНИЯ ПО ПАТТЕРНАМ")
    print("="*60)
    
    for field_num, values in sorted(field_values.items(), key=lambda x: int(x[0])):
        if not values:
            continue
            
        field_name = FIELD_MAP.get(field_num, f'Поле {field_num}')
        
        # Собрать ключевые слова
        keywords = set()
        for v in values:
            words = v.split()
            for w in words[:5]:
                w = re.sub(r'[^а-яА-Яa-zA-Z0-9]', '', w)
                if len(w) > 3:
                    keywords.add(w[:15])
        
        if keywords:
            print(f"\n{field_name}:")
            print(f"  Ключевые слова: {list(keywords)[:8]}")


if __name__ == '__main__':
    print("Сбор данных...")
    all_data = collect_all_data()
    field_values = analyze_patterns(all_data)
    generate_pattern_suggestions(field_values)