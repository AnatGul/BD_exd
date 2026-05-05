# -*- coding: utf-8 -*-
"""
Анализ Excel файлов для улучшения паттернов field_mapper
Без OCR - только на основе известных значений из переводов
"""
import os
import re
from collections import defaultdict
from openpyxl import load_workbook

# Сопоставление номеров полей Excel -> названия field_mapper
FIELD_MAP = {
    '1': 'Номер документа',
    '2': 'Наименование экспортера',
    '3': 'Год',
    '5': 'Код офиса экспорта',
    '6': 'Регистрационный номер BIN',
    '7': 'Номер коносамента',
    '8': 'Наименование получателя/грузополучателя',
    '14': 'Декларант/Агент',
    '15': 'Адрес декларанта/агента',
    '16': 'Код агента',
    '17': 'Условия поставки',
    '18': 'Код условий поставки',
    '20': 'Код валюты',
    '22': 'Общая стоимость',
    '23': 'Курс валют',
    '24': 'Страна происхождения',
    '25': 'Порт погрузки',
    '27': 'Наименование авиакомпании',
    '28': 'Код авиакомпании',
    '29': 'Наименование банка',
    '30': 'Код банка',
    '31': 'Общая декларируемая стоимость',
    '32': 'Номер места',
    '33': 'Тип упаковки',
    '34': 'Код ТН ВЭД',
    '35': 'Описание товара',
    '37': 'Количество мест (ед)',
    '38': 'Количество мест (ед.изм)',
    '41': 'Номер CRF/EXP',
    '44': 'Сектор и фонд',
    '46': 'Номер коносамента',
    '48': 'Номер документа',
}

# Собираем все значения для каждого поля
field_values = defaultdict(list)

# Все файлы xlsx в папке
xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
print(f"Найдено {len(xlsx_files)} Excel файлов")

for fname in xlsx_files:
    try:
        wb = load_workbook(fname, read_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(values_only=True):
            field_name = row[0]  # Колонка A
            value = row[2]       # Колонка C
            
            if field_name and value:
                # Извлечь номер поля
                match = re.match(r'^(\d+)\.', str(field_name))
                if match:
                    field_num = match.group(1)
                    if field_num in FIELD_MAP:
                        field_values[field_num].append(str(value))
    except Exception as e:
        print(f"Ошибка {fname}: {e}")

# Вывести все собранные значения по полям
print("\n" + "="*60)
print("СОБРАННЫЕ ЗНАЧЕНИЯ ПО ПОЛЯМ")
print("="*60)

for field_num, values in sorted(field_values.items(), key=lambda x: int(x[0])):
    field_name = FIELD_MAP.get(field_num, f"Поле {field_num}")
    print(f"\n{field_num}. {field_name} ({len(values)} значений):")
    
    # Показать уникальные значения (первые 5)
    unique_vals = list(set(values))[:5]
    for v in unique_vals:
        print(f"   - {v[:60]}")

# Создать обновленные паттерны для field_mapper
print("\n" + "="*60)
print("НОВЫЕ ПАТТЕРНЫ ДЛЯ FIELD_MAPPER")
print("="*60)

new_patterns = {}

# Для каждого поля собираем характерные паттерны
for field_num, values in field_values.items():
    field_name = FIELD_MAP.get(field_num)
    if not field_name:
        continue
    
    # Собираем общие ключевые слова из значений
    keywords = set()
    for v in values:
        words = v.split()
        for word in words[:3]:  # Первые 3 слова
            if len(word) > 3 and not word.isdigit():
                keywords.add(word[:15])
    
    if keywords:
        new_patterns[field_name] = list(keywords)[:5]

for field, patterns in new_patterns.items():
    print(f"\n{field}:")
    print(f"  Ключевые слова: {patterns}")