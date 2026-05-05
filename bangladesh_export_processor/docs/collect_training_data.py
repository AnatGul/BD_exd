# -*- coding: utf-8 -*-
"""
Сбор данных для обучения: Excel (переводы) + OCR из PDF
Создает обучающую выборку для улучшения паттернов field_mapper
"""
import os
import re
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from src.ocr_reader import OCRReader

# Сопоставление PDF ↔ Excel файлов
PDF_EXCEL_MAP = {
    'TTL-323-26 (SB).pdf': 'exp TTL-323-26.xlsx',
    'SKDL-163-2026_SB.pdf': 'exd SKDL-163-2026.xlsx',
    'SKDL-160-2026-SB.pdf': 'exp SKDL-160-2026.xlsx',
    'SKDL-147-2026 -SB.pdf': 'exd SKDL-147-2026.xlsx',
    'TTL-240-26- CO & SB-страницы-2.pdf': 'exp TTL-240-26.xlsx',
    'SKDL-97-2026 (CO+SB)-страницы-2.pdf': 'exp SKDL-97-2026.xlsx',
    'TTL-196-26- CO & SB-страницы-2.pdf': 'exp TTL-196-26.xlsx',
    'TTL-193-26- CO & SB-страницы-2.pdf': 'exp TTL-193-26.xlsx',
    'SKDL-68-2026 (CO+SB)-страницы-2.pdf': 'exp SKDL-68-2026.xlsx',
    'INV-102-2026 (CO_SB)-страницы-2.pdf': 'exp ZKLGLORIA1022026.xlsx',
    'INV-097-2026 (CO_SB)-страницы-2.pdf': 'exp ZKLGLORIA0972026.xlsx',
}

# Маппинг номеров полей Excel в названия field_mapper
FIELD_NUMBER_MAP = {
    '1': 'Номер документа',
    '2': 'Наименование экспортера',
    '3': 'Год',
    '5': 'Код офиса экспорта',
    '6': 'Регистрационный номер BIN',
    '7': 'Номер коносамента',
    '8': 'Наименование получателя/грузополучателя',
    '14': 'Декларант/Агент',
    '15': 'Адрес декларанта/агента',
    '16': 'Код агента',
    '17': 'Условия поставки',
    '18': 'Код условий поставки',
    '20': 'Код валюты',
    '22': 'Общая стоимость',
    '23': 'Курс валют',
    '24': 'Страна происхождения',
    '25': 'Порт погрузки',
    '27': 'Наименование авиакомпании',
    '28': 'Код авиакомпании',
    '29': 'Наименование банка',
    '30': 'Код банка',
    '31': 'Общая декларируемая стоимость',
    '32': 'Номер места',
    '33': 'Тип упаковки',
    '34': 'Код ТН ВЭД',
    '35': 'Описание товара',
    '37': 'Количество мест (ед)',
    '38': 'Количество мест (ед.изм)',
    '41': 'Номер CRF/EXP',
    '44': 'Сектор и фонд',
    '46': 'Номер коносамента',
    '48': 'Номер документа',
}


def read_excel_fields(excel_path: str) -> Dict[str, str]:
    """Читает Excel и возвращает {номер_поля: значение}"""
    fields = {}
    
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(values_only=True):
        field_name = row[0]  # Колонка A
        value = row[2]       # Колонка C
        
        if field_name and value:
            # Извлечь номер поля (например "2. Наименование экспортера" -> "2")
            match = re.match(r'^(\d+)\.', str(field_name))
            if match:
                field_num = match.group(1)
                fields[field_num] = str(value)
    
    return fields


def run_ocr_on_pdf(pdf_path: str) -> List[Dict]:
    """Запускает OCR на PDF и возвращает текст"""
    reader = OCRReader(use_tesseract=True, use_preprocessing=True)
    return reader.read_image(pdf_path)


def create_training_data():
    """Создает обучающую выборку"""
    training_data = []
    
    for pdf_name, excel_name in PDF_EXCEL_MAP.items():
        print(f"\n=== {excel_name} ===")
        
        # Читать Excel
        excel_path = os.path.join('.', excel_name)
        if not os.path.exists(excel_path):
            print(f"  Excel не найден: {excel_path}")
            continue
            
        excel_fields = read_excel_fields(excel_path)
        print(f"  Excel: {len(excel_fields)} полей")
        for k, v in list(excel_fields.items())[:5]:
            print(f"    {k}: {v[:40]}...")
        
        # Запустить OCR
        pdf_path = os.path.join('.', pdf_name)
        if not os.path.exists(pdf_path):
            print(f"  PDF не найден: {pdf_path}")
            continue
            
        print(f"  Running OCR on {pdf_name}...")
        ocr_results = run_ocr_on_pdf(pdf_path)
        ocr_text = " ".join([r['text'] for r in ocr_results])
        print(f"  OCR: {len(ocr_text)} chars, {len(ocr_results)} elements")
        
        training_data.append({
            'pdf': pdf_name,
            'excel': excel_name,
            'excel_fields': excel_fields,
            'ocr_text': ocr_text,
            'ocr_results': ocr_results,
        })
    
    return training_data


def analyze_patterns(training_data: List[Dict]):
    """Анализирует паттерны из обучающей выборки"""
    
    # Для каждого поля собираем все значения из Excel и ищем их в OCR
    field_patterns = {}
    
    for data in training_data:
        for field_num, value in data['excel_fields'].items():
            if field_num not in FIELD_NUMBER_MAP:
                continue
                
            field_name = FIELD_NUMBER_MAP[field_num]
            
            # Ищем это значение в OCR тексте
            ocr_lower = data['ocr_text'].lower()
            value_lower = value.lower()
            
            # Проверяем точное совпадение
            if value_lower[:20] in ocr_lower:
                if field_name not in field_patterns:
                    field_patterns[field_name] = []
                field_patterns[field_name].append(value[:50])
    
    print("\n=== Найденные паттерны ===")
    for field, values in field_patterns.items():
        print(f"\n{field}:")
        for v in set(values):
            print(f"  - {v}")


def save_training_data(training_data: List[Dict], output_path: str = 'training_data.json'):
    """Сохраняет обучающую выборку в JSON"""
    import json
    
    # Упрощаем для JSON (ocr_results слишком большой)
    simplified = []
    for data in training_data:
        simplified.append({
            'pdf': data['pdf'],
            'excel': data['excel'],
            'excel_fields': data['excel_fields'],
            'ocr_text': data['ocr_text'][:5000],  # Ограничить длину
        })
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(simplified, f, ensure_ascii=False, indent=2)
    
    print(f"\nСохранено в {output_path}")


if __name__ == '__main__':
    print("Сбор данных для обучения...")
    training_data = create_training_data()
    analyze_patterns(training_data)
    save_training_data(training_data)