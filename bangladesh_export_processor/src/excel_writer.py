# -*- coding: utf-8 -*-
"""
Excel Writer Module - Creates Excel files from translated Bangladesh Export Declarations
"""
import os
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class ExcelWriter:
    """
    Creates Excel files with translation for Bangladesh Export Declarations
    """
    
    def __init__(self):
        """Initialize Excel writer"""
        self.header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        self.header_font = Font(bold=True)
    
    def create_workbook(self) -> openpyxl.Workbook:
        """
        Create new Excel workbook
        
        Returns:
            New workbook
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Перевод'
        return wb
    
    def write_headers(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """
        Write headers to worksheet
        
        Args:
            ws: Worksheet
        """
        ws.append(['Поле', 'Исходное значение', 'Перевод'])
        
        # Format header row
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def write_fields(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                    original_fields: Dict[str, str], translated_fields: Dict[str, str]):
        """
        Write field data to worksheet
        
        Args:
            ws: Worksheet
            original_fields: Original field values
            translated_fields: Translated field values
        """
        # Write in order defined by STATIC_FIELDS from field_mapper
        from .field_mapper import FieldMapper
        
        mapper = FieldMapper()
        static_fields = mapper.STATIC_FIELDS
        
        for field_name in static_fields:
            original = original_fields.get(field_name, '')
            translated = translated_fields.get(field_name, '')
            
            if original or translated:
                ws.append([field_name, original, translated])
        
        # Add any extra fields not in static list
        for field_name, original in original_fields.items():
            if field_name not in static_fields:
                translated = translated_fields.get(field_name, '')
                ws.append([field_name, original, translated])
    
    def write(self, original_fields: Dict[str, str], translated_fields: Dict[str, str],
             output_path: str):
        """
        Write fields to Excel file
        
        Args:
            original_fields: Original field values
            translated_fields: Translated field values
            output_path: Path to output Excel file
        """
        wb = self.create_workbook()
        ws = wb.active
        
        self.write_headers(ws)
        self.write_fields(ws, original_fields, translated_fields)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 45
        
        # Save
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        
        print(f'Created: {output_path}')
    
    def write_simple(self, fields_data: List[Tuple[str, str, str]], output_path: str):
        """
        Write simple field data (field, original, translated)
        
        Args:
            fields_data: List of (field, original, translated) tuples
            output_path: Path to output Excel file
        """
        wb = self.create_workbook()
        ws = wb.active
        
        self.write_headers(ws)
        
        for field_name, original, translated in fields_data:
            ws.append([field_name, original, translated])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 45
        
        # Save
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        
        print(f'Created: {output_path}')


def test_excel_writer():
    """Test Excel writer"""
    test_fields = {
        'Номер документа': 'EXD AFL-GJ-2026-005-M',
        'Тип документа': 'EX',
        'Код таможни/тариф': 'Custom House, Dhaka',
        'Дата': '17/04/2026',
        'Наименование экспортера': 'Asdwa Fashion Ltd.',
        'Регистрационный номер BIN': '002006207-0306',
        'Адрес экспортера': 'Kandail, Satgram, Narsingdi Sadar; Madhabdi PS; Narshingdi-1603; Bangladesh',
        'Наименование получателя/грузополучателя': 'JSC "Gloria Jeans Corporation"',
        'Адрес получателя': 'Stachki avenue 184, 344090 Rostov-on-Don City, Russia',
        'Страна получателя': 'Bangladesh',
        'Код валюты': 'USD',
        'Общая стоимость': '12,320.00',
    }
    
    test_translated = {
        'Номер документа': 'EXD AFL-GJ-2026-005-M',
        'Тип документа': 'EX',
        'Код таможни/тариф': 'Таможня г. Дакка',
        'Дата': '17/04/2026',
        'Наименование экспортера': 'Асдва Фэшн Лтд.',
        'Регистрационный номер BIN': '002006207-0306',
        'Адрес экспортера': 'Бангладеш, р-н Нарсингди, г. Нарсингди-1603',
        'Наименование получателя/грузополучателя': 'АО "КОРПОРАЦИЯ ГЛОРИЯ ДЖИНС"',
        'Адрес получателя': 'пр. Стачки, 344090 г. Ростов-на-Дону, Россия',
        'Страна получателя': 'Бангладеш',
        'Код валюты': 'USD',
        'Общая стоимость': '12,320.00',
    }
    
    writer = ExcelWriter()
    writer.write(test_fields, test_translated, 'test_output.xlsx')
    
    os.remove('test_output.xlsx')
    print("Test passed!")


if __name__ == '__main__':
    test_excel_writer()